
from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성
ws.title = "Mysheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "9664c8" # 구글 RGB web에서 찾을 수 있음, RGB 형태로 값을 넣어주면 탭 색상 변경

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 두번째 index에 sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 sheet에 접근

print(wb.sheetnames) # 현재 생성되어있는 sheet의 이름을 리스트 형식으로 출력

# sheet 복사
new_ws["A1"] = "Test" # new_ws가 가리키는 NewSheet에서 A1부분에 Test값이 들어간다.
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")