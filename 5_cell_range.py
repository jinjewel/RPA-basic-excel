from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string # coordinate를 사용하기위해 선언

wb = Workbook()
ws = wb.active

ws.append(["번호","수학","영어"])
for i in range(1,11):
    ws.append([str(i)+"번", randint(0,100), randint(0,100)])

# # 영어 column만 가지고 오기
# col_C = ws["C"]
# # print(col_C)
# for cell in col_C:
#     print(cell.value)

# # 수학, 영어 column 함께 가지고 오기
# col_range = ws["B:C"]
# # print(col_C)
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# # 첫번째 row 만 가지고 오기
# row_title = ws[1]
# for cell in row_title:
#     print(cell.value, end=" ")
# print()    

# # 2~6줄의 정보 가져오기
# row_range = ws[2:6]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()    

# # 처음부터 끝까지 정보 가져오기
# row_total_range = ws[1:ws.max_row]
# for rows in row_total_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()   

# # cell의 좌표정보를 가져오기
# row_total_range = ws[1:ws.max_row]
# for rows in row_total_range:
#     for cell in rows:
#         # print(cell.coordinate, end=" ") # 셀 정보를 'A3' 처럼 가로세로가 합쳐진 형식으로 뽑아낸다.
#         xy = coordinate_from_string(cell.coordinate)
#         print(xy, end=" ") # cell 정보를 ('A', 3) 처럼 튜플로 가로 세로가 묶어진 형식으로 뽑아낸다.
#         print(xy[0], end="  ") # 튜플 정보에서 column 정보만 가져온다.
#         print(xy[1], end="  ") # 튜플 정보에서 row 정보만 가져온다.
#     print()   

# # 전체 row
# print(tuple(ws.rows))
# print()
# for row in tuple(ws.rows):
#     print(row[1].value)

# # 전체 columns
# print(tuple(ws.columns))
# print()
# for column in tuple(ws.columns):
#     print(column[0].value)

# # 전체 row
# for row in ws.iter_rows(): 
#     print(row[1].value)

# # 전체 column
# for column in ws.iter_cols():
#     print(column[0].value)

# 1번째 줄부터 5번째 줄까지 row를 반환
# for row in ws.iter_rows(min_row=1, max_row=5): 
#     print(row[1].value)

# # 1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지 row를 반환
# for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3): 
#     print(row[0].value, row[1].value) # 수학, 영어
#     print(row)

# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지 coiumn를 반환
for column in ws.iter_cols(min_row=2, max_row=11, min_col=2, max_col=3): 
    print(column[0].value,column[1].value,column[2].value) 
    print(column)






wb.save("sample.xlsx")    
