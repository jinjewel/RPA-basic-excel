from tkinter.ttk import Style
from openpyxl import load_workbook
from openpyxl.styles import Font # 폰트를 사용하기 위해 선언
from openpyxl.styles import Border, Side # 테두리를 사용하기 위해 선언
from openpyxl.styles import PatternFill # 배경색상 채우기 위해
from openpyxl.styles import Alignment

wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"] # 번호
b1 = ws["B1"] # 수학
c1 = ws["C1"] # 영어

# A 열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5

# 1 행의 높이를 50 으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color="ff0000", italic=True, bold=True) # 색상은 RGB형태로 빨간색을, 기울기와 두껍게를 적용한다.
b1.font = Font(color="cc33ff", name="Arial", strike=True) # 색상과, "Arial"이란 폰트, 취소선을 적용한다.
c1.font = Font(color="0000ff", size=20, underline="single") # 색상과, 사이즈를 20, 밑줄 적용

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 75점이 넘는 셀에 대하여 초록색으로 적용
for row in ws.rows:
    for cell in row:
        # 열과 행에 대해 중앙 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center") # center, left, right, top, bottom 사용가능

        if cell.column == 1: # A 번호 행은 제외
            continue
        elif cell.row == 1: # 이름 열은 제외
            continue
        elif int(cell.value) >= 75:
            cell.border = thin_border
        
        # cell 이 정수형 데이터이고 90점보다 높으면
        if isinstance(cell.value, int) and cell.value >= 75:
            cell.fill = PatternFill(fgColor="00ff00", fill_type="solid") # 배경을 초록색으로 변경
            cell.font = Font(color="FF0000") # 폰트 색상 변경

# 틀 고정
ws.freeze_panes = "B2" # B2 기준으로 틀 고정


wb.save("sample_styie.xlsx")