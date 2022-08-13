from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active


# 성적 입력
ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))
scores = [
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20),
]
for data in scores:
    ws.append(data)

# 오류난 퀴즈2 성적을 10점으로 고치기
col_D = ws["D"]
for cell in col_D:
    if isinstance(cell.value, int):
        cell.value = 10

# 총점 및 성적 처리
ws["H1"] = "총점"
ws["I1"] = "성적"
for idx, score in enumerate(scores, start=2):
    
    # 총점 계산
    ws["H{}".format(idx)] = "=sum(B{}:G{})".format(idx, idx) 

    # 성적 계산(총점, 출석)
    sum_sco = sum(score[1:]) - score[3] + 10 # 총점
    if sum_sco >= 90:
        ws["I{}".format(idx)] = "A"
    elif sum_sco >= 80:
        ws["I{}".format(idx)] = "B"
    elif sum_sco >= 70:
        ws["I{}".format(idx)] = "C"
    else:
        ws["I{}".format(idx)] = "D"

    if score[1] < 5:
        ws["I{}".format(idx)] = "F"

# 모든 cell 중앙 정렬
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save("scores.xlsx")    