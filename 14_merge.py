from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2") # b2부터 d2까지 병합한다.
ws["B2"] = "Merge Cell"

wb.save("sample_merge.xlsx")