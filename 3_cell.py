from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "JinjewelSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6


print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 값을 출력
print(ws["A10"].value) # 값이 없을 때 None을 출력

# row = 1,2,3,4,5 ...
# column = A(1), B(2), C(3), D(4)....
print(ws.cell(row=1, column=1)) # ws["A1"]
print(ws.cell(row=1, column=2).value) # ws["B1"].value

c = ws.cell(column=3, row=1, value=10) # ws["C1"] = 10 과 동일
print(c.value) # ws["C1"].value 와 동일

# 반복문을 이용해서 핸덤 숫자 채우기
index = 1
for x in range(1,11):
    for y in range(1,11):
        # ws.cell(column=y, row=x, value=randint(0,100))
        ws.cell(column=y, row=x, value=index)
        index += 1


wb.save("sample.xlsx")