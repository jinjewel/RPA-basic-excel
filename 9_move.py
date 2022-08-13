from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

# # 번호 수학 영어 -> 번호 (국어) 수학 영어
# ws.move_range("B1:C11", rows=0, cols=1) # ""의 내용을 rows로 0칸, cols로 1칸 옮긴다.
# ws["B1"].value = "국어" # B1셀에 '국어' 입력

ws.move_range("C1:C11", rows=5, cols=-1) # ""의 내용을 rows로 5칸, cols로 -1칸(왼쪽으로 한칸) 옮긴다.

wb.save("sample_korean.xlsx")