from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    # 번호 수학 영어
    if (row[2].value) > 60:
        print(row[0].value,"번의 점수는 ",row[2].value,"점 입니다.")

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")            
